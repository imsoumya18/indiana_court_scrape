import openpyxl

x = []

wb = openpyxl.load_workbook('ind_crawl1.xlsx')
ws = wb.worksheets[0]
print(ws.cell(1, 2).value)
for i in range(2, 99):
    x.append(str(ws.cell(i, 2).value))

print(x)
